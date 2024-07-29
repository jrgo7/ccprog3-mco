"""
Get all class names.
For each class, add a page:
    Seven columns-wide "Class: <Name>"
    Next row
    Column A: Method
    Column B: #
    Column C: Test Description
    Column D: Sample Input Data
    Column E: Expected Output
    Column F: Actual Output
    Column G: P/F
    Next row: do the ff. pattern for each method name
        Method name 1
                    2
                    3
"""
import os
import regex
from openpyxl import Workbook

FOLDER_PATH: str = "./src" # Change this

""" Retrieved from https://stackoverflow.com/a/77307448"""
JAVA_METHOD_REGEX: regex.Pattern  = regex.compile(r"""\b(public|private|protected)\s*(<[^>]*>)?\s*(abstract|static|final)?\s*(abstract|static|final)?\s*(abstract|static|final)?\s+[^*](\w+(\.\w+)*)*(\s*<[^>]*>)?(\s*\[[^\]]*\])*(\s+\w+\s*\([^)]*\)(\s+throws\s+\w+(\s*,\s*\w+)*)?)+""")
wb = Workbook()
wb.remove(wb.active) # Remove default Sheet

def get_tab_color(file_path):
    if "controller" in file_path:
        return "8AEA92"
    elif "model" in file_path:
        return "80ADA0"
    elif "view" in file_path:
        return "5F5566"
    else:
        return "33202A"

def deal_with(file_path: str):
    folder_path, class_name = os.path.split(file_path)
    
    class_name = class_name.split(".")[0]
    wb.create_sheet(class_name)
    ws = wb[class_name]
    ws.sheet_properties.tabColor = get_tab_color(file_path)
    
    ws.append([class_name])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=7)
    ws.append(["Method", "#", "Test Description", "Sample Input Data", "Expected Output", "Actual Output", "P/F"])
        
    os.system("javap %s > \"%s\"" % (file_path, os.path.join(folder_path, class_name + "-charlotte.txt")))
    
    with open(os.path.join(folder_path, class_name + "-charlotte.txt"), "r") as file:
        listing = file.read()

    methods_matches = regex.findall(JAVA_METHOD_REGEX, listing)
    methods = [method[9].strip() for method in methods_matches]
    
    for i, method in enumerate(methods, start=1):
        print(method, end=" -> ")
        if "." in method:
            method_name = method.split(".")[0].split("(")[0] + "("
            if "ArrayList" not in method:
                method_name += method.split(".")[-1]
            else:
                method_name += "ArrayList<" + method.split(".")[-1]
        else:
            method_name = method
        print(method_name)
        ws.append([method_name, "1", "", "", "", "", "P"])
        if "get" not in method_name:
            for i in range (2, 4):
                ws.append(["", str(i), "", "", "", "", "P"])
            
    os.remove(os.path.join(folder_path, class_name + "-charlotte.txt"))
    
def main():
    method_count = 0
    out = ""
    for path, folders, files in os.walk(FOLDER_PATH):
        for folder in folders:
            for file in os.listdir(os.path.join(path, folder)):
                if file.endswith(".class"):
                    deal_with(os.path.join(path, folder, file))
    wb.save("Test Script Template.xlsx")
    return

if __name__ == '__main__':
    main()